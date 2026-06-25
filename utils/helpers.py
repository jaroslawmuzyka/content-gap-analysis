from io import BytesIO
import pandas as pd

def normalize_url(url):
    if not isinstance(url, str):
        return ""
    url = url.replace("https://", "").replace("http://", "").replace("www.", "")
    if url.endswith("/"):
        url = url[:-1]
    return url.strip()

def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
    return output.getvalue()

import streamlit as st

PRICING = {
    "gpt-4o-mini": {"input": 0.15, "output": 0.60},
    "gpt-5.4-mini": {"input": 0.15, "output": 0.60},
    "gpt-4o": {"input": 5.00, "output": 15.00},
    "gpt-5.5": {"input": 5.00, "output": 15.00},
    "o1-mini": {"input": 3.00, "output": 12.00},
    "gpt-4-turbo": {"input": 10.00, "output": 30.00},
    "gpt-3.5-turbo": {"input": 0.50, "output": 1.50},
    "o3-mini": {"input": 3.00, "output": 12.00},
}

def track_usage(model_name, prompt_tokens, completion_tokens):
    if "total_api_cost" not in st.session_state:
        st.session_state.total_api_cost = 0.0
    if "total_tokens" not in st.session_state:
        st.session_state.total_tokens = {"prompt": 0, "completion": 0}
        
    rates = PRICING.get(model_name, PRICING["gpt-4o-mini"])
    cost = (prompt_tokens / 1000000.0) * rates["input"] + (completion_tokens / 1000000.0) * rates["output"]
    
    st.session_state.total_api_cost += cost
    st.session_state.total_tokens["prompt"] += prompt_tokens
    st.session_state.total_tokens["completion"] += completion_tokens

def render_wow_metrics():
    if "global_stats" not in st.session_state:
        return
        
    st.markdown("### 📊 Aktualne Statystyki Analizy (WOW Metrics)")
    stats = st.session_state.global_stats
    
    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Frazy", stats.get("przeanalizowane_frazy", 0))
    col2.metric("Strony Konkurencji", stats.get("strony_konkurencji", 0))
    col3.metric("Strony Własne", stats.get("strony_wlasne", 0))
    col4.metric("Klastry/Pomysły", stats.get("wygenerowane_pomysly", 0))
    col5.metric("Zidentyfikowane Luki", stats.get("content_gaps", 0))
    st.markdown("---")

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

def to_excel_multi(sheets_dict):
    output = BytesIO()
    has_sheets = False
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in sheets_dict.items():
            if df is not None and not df.empty:
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                has_sheets = True
        
        if not has_sheets:
            pd.DataFrame({"Informacja": ["Brak danych (puste tabele)"]}).to_excel(writer, index=False, sheet_name="Brak Danych")
                
        # Access the workbook to apply formatting
        workbook = writer.book
        
        # Define some basic styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Format header row
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                
            # Auto-adjust columns width based on data
            for col_idx, col in enumerate(worksheet.columns, 1):
                max_length = 0
                column = get_column_letter(col_idx)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                # Set a little padding, max width 60, min width 10
                adjusted_width = max(10, min(max_length + 2, 60))
                worksheet.column_dimensions[column].width = adjusted_width

    return output.getvalue()

def clean_json(text):
    if not isinstance(text, str): return ""
    text = text.strip()
    if text.startswith('```json'):
        text = text[7:]
    if text.startswith('```'):
        text = text[3:]
    if text.endswith('```'):
        text = text[:-3]
    text = text.strip()
    
    # Próba prostego zamknięcia urwanego JSON-a, jeśli brakuje nawiasów
    import json
    try:
        json.loads(text)
        return text
    except json.JSONDecodeError:
        # Bardzo prosta heurystyka naprawy uciętych stringów i nawiasów
        fixed = text.rstrip()
        open_brackets = []
        in_string = False
        escape = False
        for char in fixed:
            if escape:
                escape = False
                continue
            if char == '"':
                in_string = not in_string
            elif char == '\\':
                escape = True
            elif not in_string:
                if char in '{[':
                    open_brackets.append(char)
                elif char in '}]':
                    if open_brackets:
                        open_brackets.pop()
        
        if in_string:
            fixed += '"'
        
        while open_brackets:
            bracket = open_brackets.pop()
            if bracket == '{':
                fixed += '}'
            elif bracket == '[':
                fixed += ']'
                
        return fixed
def get_step2_excel_sheets(product_analysis_list):
    p1_fakty_data = []
    p2_zastosowania_data = []
    p2_strategia_data = []
    p3_frazy_data = []
    p4_frazy_data = []
    p5_kontekst_data = []
    
    for item in product_analysis_list:
        url = item["url"]
        j1 = item.get("json1", {})
        j2 = item.get("json2", {})
        j3 = item.get("json3", {})
        j4 = item.get("json4", {})
        ctx = item.get("products_context", "")
        
        # P1 - Fakty ze strony
        sklad_dict = j1.get("sklad") or {}
        skladniki = sklad_dict.get("skladniki_aktywne_lub_kluczowe") or [] if isinstance(sklad_dict, dict) else []
        skladniki_str = ", ".join([s.get("skladnik", "") for s in skladniki if isinstance(s, dict)])
        
        wskazania = j1.get("wskazania_i_zastosowania") or []
        wskazania_str = ", ".join([w.get("nazwa", "") for w in wskazania if isinstance(w, dict)])
        
        dzialanie = j1.get("dzialanie_i_mechanizm") or []
        dzialanie_str = ", ".join([d.get("dzialanie", "") for d in dzialanie if isinstance(d, dict)])
        
        grupy = j1.get("grupy_docelowe_wprost") or []
        grupy_str = ", ".join([g.get("grupa", "") for g in grupy if isinstance(g, dict)])
        
        p1 = j1.get("produkt") or {}
        if not isinstance(p1, dict): p1 = {}
        
        p1_fakty_data.append({
            "URL": url,
            "Nazwa": p1.get("nazwa", ""),
            "Status": p1.get("status_produktu", ""),
            "Kategoria": p1.get("kategoria", ""),
            "Postać": p1.get("postac", ""),
            "Dostępny bez recepty": str(p1.get("czy_dostepny_bez_recepty", "")),
            "Składniki Aktywne": skladniki_str,
            "Wskazania wprost": wskazania_str,
            "Działanie/Mechanizm": dzialanie_str,
            "Grupy Docelowe": grupy_str
        })
        
        # P2 - Analiza Zastosowań (Nowa struktura: Insighty)
        insighty = j2.get("najwazniejsze_insighty") or []
        if not isinstance(insighty, list): insighty = []
        for zast in insighty:
            if isinstance(zast, dict):
                p2_zastosowania_data.append({
                    "URL": url,
                    "Insight": zast.get("insight", ""),
                    "Związek": zast.get("zwiazek_przyczynowo_skutkowy", ""),
                    "Problem Użytkownika": zast.get("problem_uzytkownika", ""),
                    "Rola Produktu": zast.get("rola_produktu", ""),
                    "Bezpieczny Kierunek": zast.get("bezpieczny_kierunek_contentu", ""),
                    "Dopasowanie": zast.get("dopasowanie_do_produktu", ""),
                    "Wymaga Weryfikacji": str(zast.get("wymaga_weryfikacji", ""))
                })
        
        # P2 - Profil Strategiczny (Nowa struktura: Podsumowanie)
        podsumowanie = j2.get("podsumowanie") or {}
        if not isinstance(podsumowanie, dict): podsumowanie = {}
        
        p2_strategia_data.append({
            "URL": url,
            "Najlepszy Punkt Zaczepienia": podsumowanie.get("najlepszy_punkt_zaczepienia", ""),
            "Największa Szansa Contentowa": podsumowanie.get("najwieksza_szansa_contentowa", ""),
            "Największe Ryzyko Nadużycia": podsumowanie.get("najwieksze_ryzyko_naduzycia", "")
        })
        
        # P3 - Frazy z Faktów
        seed_3 = j3.get("seed_keywords") or []
        if not isinstance(seed_3, list): seed_3 = []
        for k in seed_3:
            p3_frazy_data.append({
                "URL": url,
                "Fraza (Seed Keyword)": str(k).strip()
            })
            
        # P4 - Frazy z Analizy
        seed_4 = j4.get("seed_keywords") or []
        if not isinstance(seed_4, list): seed_4 = []
        for k in seed_4:
            p4_frazy_data.append({
                "URL": url,
                "Fraza (Seed Keyword)": str(k).strip()
            })
            
        # P5 - Kontekst Content Gap
        p5_kontekst_data.append({
            "URL": url,
            "Kontekst Produktu (Brief)": ctx
        })
    
    excel_sheets = {
        "P1 - Fakty ze strony": pd.DataFrame(p1_fakty_data),
        "P2 - Analiza Zastosowań": pd.DataFrame(p2_zastosowania_data),
        "P2 - Profil Strategiczny": pd.DataFrame(p2_strategia_data),
        "P3 - Frazy z Faktów": pd.DataFrame(p3_frazy_data),
        "P4 - Frazy z Analizy": pd.DataFrame(p4_frazy_data),
        "P5 - Kontekst CG": pd.DataFrame(p5_kontekst_data)
    }
    return excel_sheets
